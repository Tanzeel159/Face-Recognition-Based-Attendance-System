import cognitive_face as CF
import global_variables as global_var
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import Cell
import time
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)



#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports.xlsx")
sheet = wb['Cse15']

def getDateColumn():
	for i in range(1, len(list(sheet.rows)[0]) + 1):
		col = get_column_letter(i)
		if sheet['%s%s'% (col,'1')].value == currentDate:
			return col

Key = global_var.key

CF.Key.set(Key)

BASE_URL = global_var.BASE_URL  # Replace with your regional Base URL
CF.BaseUrl.set(BASE_URL)

connect = sqlite3.connect("Face-DataBase")
#c = connect.cursor()

attend = [0 for i in range(60)]	

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'Cropped_faces')
for filename in os.listdir(directory):
	if filename.endswith(".jpg"):
		imgurl = urllib.request.pathname2url(os.path.join(directory, filename))
		imgurl = imgurl[3:]
#		print("imgurl = {}".format(imgurl))
		res = CF.face.detect(imgurl)
		print("Res = {}".format(res))

		if len(res) < 1:
			print("No face detected.")
			continue
			
		faceIds = []
		for face in res:
			faceIds.append(face['faceId'])
		res = CF.face.identify(faceIds, global_var.personGroupId)
		print(filename)
		print("res = {}".format(res))

		for face  in res:
			if not face['candidates']:
				print("Unknown")
			else:
				personId = face['candidates'][0]['personId']
				print("personid = {}".format(personId))
				#cmd =  + personId
				cur = connect.execute("SELECT * FROM Students WHERE personID = (?)", (personId,))
				#print("cur = {}".format(cur))
				for row in cur:
					print("aya")
					print("row = {}".format(row))
					attend[int(row[0])] += 1
				print("---------- " + row[1] + " recognized ----------")
		time.sleep(6)
		
for row in range(2, len(list(sheet.columns)[0]) + 1):
	rn = sheet.cell(row = row, column  =1).value
	if rn is not None:
		print("rn = {}".format(rn))
		rn = rn[-2:]
		if attend[int(rn)] != 0:
			col = getDateColumn()
			print("col = {}".format(col))
			sheet['%s%s' % (col, str(row))] = 1

wb.save(filename = "reports.xlsx")	 	
#currentDir = os.path.dirname(os.path.abspath(__file__))
#imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(imgurl)
#faceIds = []
#for face in res:
 #   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
#print res
